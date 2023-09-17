import random
import openpyxl as xl
import math as m
import os

# Get directory
script_dir = os.path.dirname(os.path.abspath(__file__))


class Node:
    def __init__(self, data):
        self.data = data
        self.children = []

def build_tree(data, depth, max_depth):
    if depth >= max_depth or not data:
        return None

    root_data = random.choice(data)
    root = Node(root_data)
    data.remove(root_data)

    for _ in range(len(data)):
        child = build_tree(data, depth + 1, max_depth)
        if child:
            root.children.append(child)

    return root

def depth_limited_search(node, goal_node, current_depth, depth_limit, track):
    if node.data == goal_node:
        track.append(node.data)
        track.sort()
        return track
    
    track.append(node.data)
    
    for child in node.children:
        result = depth_limited_search(child,  goal_node, current_depth + 1, depth_limit, track)
        if result: 
            track.sort()
            return track

def euclidean(P1_0_Letter, P1_1_Letter, P2_0_Letter, P2_1_Letter, sheetname):
    
    # All possible extensions
    extensions = {'xlsx':'xlsx', 'xlsm':'xlsm', 'xltx':'xltx', 'xltm':'xltm'}
    
    # Iterate through directory to find the file
    # And load the file if found
    for extension in extensions.values():
        try:
            workbook = xl.load_workbook(os.path.join(script_dir, f"{sheetname}.{extension}"))
            sheet = workbook.active
        except FileNotFoundError:
            continue
    
    # Initialize lists of Pointer 1, Element 0 and Element 1 Data
    P1_0_Data = []
    P1_1_Data = []
    
    # Starting row to get the data. 2, because 1 is usually names
    start_row = 2
    
    # Key as name of garden & value as the distance
    labels = dict()
    
    # Assigning names to keys to dictionary
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'A{row}']
        if cell.value is not None:
            labels[cell.value] = 0  # Append the name as a key
        else: 
            break # Stop the loop when an empty cell is encountered
            
    # P1 Data, Element 0
    for row in range(start_row, sheet.max_row + 1):  
            cell = sheet[f'{P1_0_Letter}{row}']
            forlabel = sheet[f'A{row}']
            if cell.value is not None:
                P1_0_Data.append(cell.value)  # Append the value to the list
                labels[forlabel.value] = cell.value # Assign the label its value
            else:
                break # Stop the loop when an empty cell is encountered
        
    # P1 Data, Element 1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P1_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P1_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    solve = []
    
    # Check if P2_0 and P2_1 parameters is a number
    if (isinstance(P2_0_Letter, float) or isinstance(P2_0_Letter, int)) and (isinstance(P2_1_Letter, float) or isinstance(P2_1_Letter, int)):
        
        # If both parameters are numbers, solve euclidean
        for row in range(start_row, sheet.max_row + 1):
            for i in range(len(P1_0_Data)):
                D = m.sqrt(((P1_0_Data[i] - P2_0_Letter) ** 2) + ((P1_1_Data[i] - P2_1_Letter) ** 2))
                solve.append(D) # Append result in a list
                labels[sheet[f'A{row}'].value] = D # Assign result to specified key
                row += 1
                
            workbook.close()
            break
        return solve, labels
        
    # Pointer 2, Element 0 and Element 1 initialize as lists
    P2_0_Data = []
    P2_1_Data = []
        
    # Getting the data for Pointer 2, E0
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_0_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_0_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    # Getting the data for Pointer 2, E1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
     
    # Solve each row to euclidean
    for row in range(start_row, sheet.max_row + 1):  
        for i in range(len(P1_0_Data)):
            D = m.sqrt(((P1_0_Data[i] - P2_0_Data[i]) ** 2) + ((P1_1_Data[i] - P2_1_Data[i]) ** 2))
            solve.append(D) # Append result in a list
            labels[sheet[f'A{row}'].value] = D # Assign result to specified key
            row += 1
        workbook.close()
        break
    return solve, labels

def manhattan(P1_0_Letter, P1_1_Letter, P2_0_Letter, P2_1_Letter, sheetname):
    # All possible extensions
    extensions = {'xlsx':'xlsx', 'xlsm':'xlsm', 'xltx':'xltx', 'xltm':'xltm'}
    
    # Iterate through directory to find the file
    # And load the file if found
    for extension in extensions.values():
        try:
            workbook = xl.load_workbook(os.path.join(script_dir, f"{sheetname}.{extension}"))
            sheet = workbook.active
        except FileNotFoundError:
            continue
    
    # Initialize lists of Pointer 1, Element 0 and Element 1 Data
    P1_0_Data = []
    P1_1_Data = []
    
    # Starting row to get the data. 2, because 1 is usually names
    start_row = 2
    
    # Key as name of garden & value as the distance
    labels = dict()
    
    # Assigning names to keys to dictionary
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'A{row}']
        if cell.value is not None:
            labels[cell.value] = 0  # Append the name as a key
        else: 
            break # Stop the loop when an empty cell is encountered
            
    # P1 Data, Element 0
    for row in range(start_row, sheet.max_row + 1):  
            cell = sheet[f'{P1_0_Letter}{row}']
            forlabel = sheet[f'A{row}']
            if cell.value is not None:
                P1_0_Data.append(cell.value)  # Append the value to the list
                labels[forlabel.value] = cell.value # Assign the label its value
            else:
                break # Stop the loop when an empty cell is encountered
        
    # P1 Data, Element 1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P1_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P1_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    solve = []
    
    # Check if P2_0 and P2_1 parameters is a number
    if (isinstance(P2_0_Letter, float) or isinstance(P2_0_Letter, int)) and (isinstance(P2_1_Letter, float) or isinstance(P2_1_Letter, int)):
        
        # If both parameters are numbers, solve euclidean
        for row in range(start_row, sheet.max_row + 1):
            for i in range(len(P1_0_Data)):
                D = (P1_0_Data[i] - P2_0_Letter) + (P1_1_Data[i] - P2_1_Letter)
                solve.append(D) # Append result in a list
                labels[sheet[f'A{row}'].value] = D # Assign result to specified key
                row += 1
            workbook.close()
            break
        return solve, labels
        
    # Pointer 2, Element 0 and Element 1 initialize as lists
    P2_0_Data = []
    P2_1_Data = []
        
    # Getting the data for Pointer 2, E0
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_0_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_0_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    # Getting the data for Pointer 2, E1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
     
    # Solve each row to euclidean
    for row in range(start_row, sheet.max_row + 1):
        for i in range(len(P1_0_Data)):
            D = (P1_0_Data[i] - P2_0_Data[i]) + (P1_1_Data[i] - P2_1_Data[i])
            solve.append(D) # Append result in a list
            labels[sheet[f'A{row}'].value] = D # Assign result to specified key
            row += 1
        workbook.close()
        break
    return solve, labels

def minkowski(P1_0_Letter, P1_1_Letter, P2_0_Letter, P2_1_Letter, sheetname, p=3):
    
    # All possible extensions
    extensions = {'xlsx':'xlsx', 'xlsm':'xlsm', 'xltx':'xltx', 'xltm':'xltm'}
    
    # Iterate through directory to find the file
    # And load the file if found
    for extension in extensions.values():
        try:
            workbook = xl.load_workbook(os.path.join(script_dir, f"{sheetname}.{extension}"))
            sheet = workbook.active
        except FileNotFoundError:
            continue
    
    # Initialize lists of Pointer 1, Element 0 and Element 1 Data
    P1_0_Data = []
    P1_1_Data = []
    
    # Starting row to get the data. 2, because 1 is usually names
    start_row = 2
    
    # Key as name of garden & value as the distance
    labels = dict()
    
    # Assigning names to keys to dictionary
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'A{row}']
        if cell.value is not None:
            labels[cell.value] = 0  # Append the name as a key
        else: 
            break # Stop the loop when an empty cell is encountered
            
    # P1 Data, Element 0
    for row in range(start_row, sheet.max_row + 1):  
            cell = sheet[f'{P1_0_Letter}{row}']
            forlabel = sheet[f'A{row}']
            if cell.value is not None:
                P1_0_Data.append(cell.value)  # Append the value to the list
                labels[forlabel.value] = cell.value # Assign the label its value
            else:
                break # Stop the loop when an empty cell is encountered
        
    # P1 Data, Element 1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P1_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P1_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    solve = []
    
    # Check if P2_0 and P2_1 parameters is a number
    if (isinstance(P2_0_Letter, float) or isinstance(P2_0_Letter, int)) and (isinstance(P2_1_Letter, float) or isinstance(P2_1_Letter, int)):
        
        # If both parameters are numbers, solve euclidean
        for row in range(start_row, sheet.max_row + 1):
            for i in range(len(P1_0_Data)):
                D = ((abs(P1_0_Data[i] - P2_0_Data[i]) ** p) + (abs(P1_1_Data[i] - P2_1_Data[i]) ** p)) ** (1/p)
                solve.append(D) # Append result in a list
                labels[sheet[f'A{row}'].value] = D # Assign result to specified key
                row += 1
            workbook.close()
            break
        return solve, labels
        
    # Pointer 2, Element 0 and Element 1 initialize as lists
    P2_0_Data = []
    P2_1_Data = []
        
    # Getting the data for Pointer 2, E0
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_0_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_0_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
        
    # Getting the data for Pointer 2, E1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value # Assign the label its value
        else:
            break # Stop the loop when an empty cell is encountered
     
    # Solve each row to euclidean
    for row in range(start_row, sheet.max_row + 1):
        for i in range(len(P1_0_Data)):
            D = ((abs(P1_0_Data[i] - P2_0_Data[i]) ** p) + (abs(P1_1_Data[i] - P2_1_Data[i]) ** p)) ** (1/p)
            solve.append(D) # Append result in a list
            labels[sheet[f'A{row}'].value] = D # Assign result to specified key
            row += 1
        workbook.close()
        break
    return solve, labels

data, labels = euclidean('B', 'C', 'D', 'E', 'data')
# Solve eucledian distance by: Point1(B,C) and Point2(D,E)

# data, labels = manhattan('B', 'C', 'D', 'E', 'data')
# Solve manhattan distance by: Point1(B,C) and Point2(D,E)

# data, labels = minkowski('B', 'C', 'D', 'E', 'data', p=3)
# Solve minkowski distance by: Point1(B,C) and Point2(D,E)

goal_list = list(data) # Making a copy of the list of data

# Sorting goal list and value of dictionary to ascending order
sorted_labels = dict(sorted(labels.items(), key=lambda item: item[1]))
goal_list.sort()

# Make a tree for the DLS algorithm search
root_node = build_tree(data, 0, len(labels)+1)

# Find the lowest values by using DLS
ranking = []
track = []
for i in range(len(labels)):
    ranking.append(depth_limited_search(root_node, goal_list[i], 0, len(labels)+1, track))

# Remove any sublist and append all of the values in one list
ranking_ = []
for sublist in ranking:
    if isinstance(sublist, list):
        for num in sublist:
            ranking_.append(num)
            continue
    ranking_.append(num)
        
# Converting ranking into set, to remove repetitions
store = set()
ranking__ = [x for x in ranking_ if x not in store and not store.add(x)]


# Printing the ranks of euclidean distance
print(f"Ranking of DLS: {ranking__[:3]}")
print("\nActual Ranking: ")
i = int(0)
for name, val in sorted_labels.items():
    print(f"{name} : {val}")
    i += int(1)
    if i >= 3:
        break

# Testing and Training
import pandas as pd
import numpy as np
from collections import Counter
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
import matplotlib.pyplot as plt

filename = 'data_2'
file_extensions = {'xlsx': pd.read_excel, 'csv': pd.read_csv, 'xls': pd.read_excel}

for ext, func in file_extensions.items():
    try:
        df = func(os.path.join(script_dir, f'{filename}.{ext}'))
        break
    except FileNotFoundError:
        continue

# Now that df is defined, you can modify it
# NY Garden = 1
df.loc[0, 'Name'] = 1

# Brooklyn Garden = 2
df.loc[1, 'Name'] = 2

# Queens Garden = 3
df.loc[2, 'Name'] = 3

# Snug Garden = 4
df.loc[3, 'Name'] = 4



#Rename Columns

df.rename(columns= {'LATITUDE': 'LT'}, inplace= True)
df.rename(columns= {'LONGITUDE': 'LO'}, inplace= True)
df.rename(columns= {'My Latitude': 'm_LT'}, inplace= True)
df.rename(columns= {'My Longitude': 'm_LO'}, inplace= True)
df.rename(columns= {'NAME': 'Garden Name'}, inplace= True)

df = df[['Garden Name', 'LT','LO','m_LT','m_LO']]

#Data Splitting
train = df.iloc[0:33] # NY and BRK for training
test = df.iloc[34:70] # Q and SNG for test

def euclidean_distance(row1, row2, columns, label):
    distance = 0.0

    for column in columns:
        #Only euclidean distance for features is calculated
        if column != label:
            distance += (row1[column] - row2[column])**2

    return m.sqrt(distance)


def manhattan_distance(row1, row2, columns, label):
    distance = 0.0

    for column in columns:
        if column != label:
            distance += abs(row1[column] - row2[column])

    return distance

def minkowski_distance(row1, row2, columns, label, p):

    distance = 0

    for column in columns:
        if column != label:
            distance += abs(row1[column] - row2[column]) ** p

    return distance ** (1 / p)

def KNN(train, test_row, k, label):
    temp = train.copy()

    #Calculate distance for each instance in train to single test instance
    temp['dist'] = temp.apply(lambda row: euclidean_distance(row, test_row, train.columns, label), axis=1)

    # Manhattan distance
    # temp['dist'] = temp.apply(lambda row: manhattan_distance(row, test_row, train.columns, label), axis=1)
    
    # Minkowski distance
    # temp['dist'] = temp.apply(lambda row: minkowski_distance(row, test_row, train.columns, label, 2), axis=1)

    #Getting the k neighbors having minimum distances
    sorted_distances = temp['dist'].sort_values()
    k_neighbors_distances = sorted_distances [:k]
    

    #Getting the majority label from the k neighbors
    k_neighbors = temp[temp.index.isin(k_neighbors_distances.index)]
    k_neighbors_labels = list(k_neighbors[label])
    count_labels = Counter(k_neighbors_labels)
    predicted_label = count_labels.most_common()[0][0]
    return predicted_label

#Varrying K

pred = []
actual = []
scores = []

#For Checking Values
pred_values = []
actual_values = []

for k in range (1,3):
    for i in range (test.shape[0]):
        pred.append( KNN(train=df, test_row=test.iloc[i], k=k, label = 'Garden Name')) #.iloc is the index locator,
        actual.append(test.iloc[i, 0]) #Get 0 column of i row, 0 column is the '_species_type'
        scores.append(accuracy_score(actual, pred))

        pred_values.append(KNN(train=df, test_row=test.iloc[i], k=k, label = 'Garden Name'))
        actual_values.append(test.iloc[i, 0])

    pred = []
    actual = []
    
print(f'\nScores : {scores}')
print('Mean Accuracy: %.2f%%' % (sum(scores)/float(len(scores))))

plt.plot(scores)
plt.title('Accuracy score of different k neighbors')
plt.xlabel('k neighbors')
plt.ylabel('accuracy score')
plt.show()
