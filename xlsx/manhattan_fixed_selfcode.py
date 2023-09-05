import openpyxl as xl
import math as m
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
sheet = workbook.active

first_data = []
second_data = []

third_data = []
fourth_data = []

start_row = 2  # 2 because 1 is usually labels
column_letter = input("First column for first data: ") # e.g. A

# Iterate through the column to get the first data
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        first_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
column_letter = input("Second column for Second data: ") # e.g. B
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        second_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
column_letter = input("Third column for third data: ") # e.g. B
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        third_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
column_letter = input("Fourth column for fourth data: ") # e.g. B
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        fourth_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
workbook.close()

solve = []
for i in range(len(first_data)):
    dt1 = first_data[i] - third_data[i]
    dt2 = second_data[i] - fourth_data[i]
    solve.append(dt1 + dt2)

solve.sort()
print(solve[:3])
    

