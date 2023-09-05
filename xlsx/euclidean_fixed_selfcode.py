import openpyxl as xl
import math as m
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
sheet = workbook.active

# Under classifying Unlabeled Data
classified_data_1 = 5.5 # float(input("Classified data 1: ")) # e.g. sepal length
classified_data_2 = 5 # float(input("Classified data 2: ")) # e.g. sepal width

first_data = []
second_data = []

start_row = 2  # 2 because 1 is usually labels
column_letter = input("First column for first data: ") # e.g. A

# Iterate through the column to get the first data
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        first_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
column_letter = input("Second column for Second data: ") # e.g. B
for row in range(start_row, sheet.max_row + 1):  
    cell = sheet[f'{column_letter}{row}']
    if cell.value is not None:
        second_data.append(cell.value)  # Append the value to the list
    else:
        # Stop the loop when an empty cell is encountered
        break
    
workbook.close()

solve = []
# (m.sqrt((first_data[i] - classified_data_1)**2) + ((second_data[i] - classified_data_2)**2)

for i in range(len(first_data)):
    dt1 = first_data[i] - classified_data_1
    dt2 = second_data[i] - classified_data_2
    dt1 = pow(dt1, 2)
    dt2 = pow(dt2, 2)
    solve.append(m.sqrt(dt1 + dt2))

solve.sort()
print(solve[:3])
    

