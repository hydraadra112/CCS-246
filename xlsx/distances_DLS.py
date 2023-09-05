import random
import openpyxl as xl
import math as m
import os

class Node:
    def __init__(self, data):
        self.data = data
        self.children = []
def build_random_tree(values, depth, max_depth):
    if depth >= max_depth or not values:
        return None

    root_data = random.choice(values)
    root = Node(root_data)
    values.remove(root_data)

    for _ in range(random.randint(0, len(values))):  # Randomly choose the number of children
        child = build_random_tree(values, depth + 1, max_depth)
        if child:
            root.children.append(child)

    return root

def depth_limited_search(node, goal_node, current_depth, depth_limit, visited, track):
    if node.data == goal_node:
        track.append(node.data)
        track.sort
        return track[:5]
    
    if node in visited:
        return None
    
    track.append(node.data)
    visited.add(node)
    
    for child in node.children:
        result = depth_limited_search(child, goal_node, current_depth + 1, depth_limit, visited, track)
        if result:
            track.sort()
            return track[:5]
    
    return None

def euclidean(classified_data_1, classified_data_2, column_letter_1, column_letter_2):
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
    sheet = workbook.active

    # Under classifying Unlabeled Data
    # classified_data_1 = 5.5 # float(input("Classified data 1: ")) # e.g. sepal length
    # classified_data_2 = 5 # float(input("Classified data 2: ")) # e.g. sepal width

    first_data = []
    second_data = []

    start_row = 2  # 2 because 1 is usually labels

    # Iterate through the column to get the first data
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_1}{row}']
        if cell.value is not None:
            first_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
        
    # input("Second column for Second data: ") # e.g. B
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_2}{row}']
        if cell.value is not None:
            second_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
    
    workbook.close()

    solve = []
    for i in range(len(first_data)):
        dt1 = first_data[i] - classified_data_1
        dt2 = second_data[i] - classified_data_2
        dt1 = pow(dt1, 2)
        dt2 = pow(dt2, 2)
        solve.append(m.sqrt(dt1 + dt2))
    
    return solve

def manhattan(column_letter_1, column_letter_2):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
    sheet = workbook.active

    first_data = []
    second_data = []

    start_row = 2  # 2 because 1 is usually labels

    # Iterate through the column to get the first data
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_1}{row}']
        if cell.value is not None:
            first_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
        
    # input("Second column for Second data: ") # e.g. B
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_2}{row}']
        if cell.value is not None:
            second_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
    
    workbook.close()

    solve = []
    for i in range(len(first_data)):
        solve.append(first_data[i] - second_data[i])
    
    return solve

def minkowski(column_letter_1, column_letter_2, p):
    if p <= 0:
        raise ValueError("Parameter p must be greater than 0")
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
    sheet = workbook.active

    first_data = []
    second_data = []

    start_row = 2  # 2 because 1 is usually labels

    # Iterate through the column to get the first data
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f'{column_letter_1}{row}']
        if cell.value is not None:
            first_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break

    # Iterate through the second column for Second data
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f'{column_letter_2}{row}']
        if cell.value is not None:
            second_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break

    if len(first_data) != len(second_data):
        raise ValueError("Points must have the same number of elements/dimensions")

    workbook.close()
    
    distance = 0.0
    for i in range(len(first_data)):
        distance += abs(first_data[i] - second_data[i]) ** p
    return distance ** (1 / p)

data = euclidean(5.5, 5, 'A', 'B')
# data = manhattan('A', 'C')
# minkowski = minkowski('A','B', 2)

if minkowski:
    print(f"Minkowski distance: {minkowski}")
    exit()


goal_list = list(data)
goal_list.sort()

# Define the maximum depth for the tree
max_depth = 5 
# Build a random tree from the taken data
root_node = build_random_tree(data, 0, max_depth)

# Find the 3 lowest value using DLS
ranking = []
track = []
visited = set()

for i in range(3):
    ranking.append(depth_limited_search(root_node, goal_list[i], 0, 5, visited, track))

print(f"Ranking of DLS: {ranking}")
print(f"Actual Ranking: {goal_list[:5]}")


