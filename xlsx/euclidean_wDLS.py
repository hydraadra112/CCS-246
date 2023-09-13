import random
import openpyxl as xl
import math as m
import os

class Node:
    def __init__(self, data):
        self.data = data
        self.children = []

def build_tree(data, depth, max_depth):
    if depth >= max_depth or not data:
        return None

    root_data = random.choice(data)
    root = Node(root_data)
    data.remove(root_data)

    for _ in range(len(data)):
        child = build_tree(data, depth + 1, max_depth)
        if child:
            root.children.append(child)

    return root

def depth_limited_search(node, goal_node, current_depth, depth_limit, track):
    if node.data == goal_node:
        track.append(node.data)
        track.sort()
        return track
    
    track.append(node.data)
    
    for child in node.children:
        result = depth_limited_search(child,  goal_node, current_depth + 1, depth_limit, track)
        if result: 
            track.sort()
            return track

def euclidean(P1_0_Letter, P1_1_Letter, P2_0_Letter, P2_1_Letter, sheetname):
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    extensions = {'xlsx':'xlsx', 'xlsm':'xlsm', 'xltx':'xltx', 'xltm':'xltm'}
    
    for extension in extensions.values():
        try:
            workbook = xl.load_workbook(os.path.join(script_dir, f"{sheetname}.{extension}"))
            sheet = workbook.active
        except FileNotFoundError:
            continue
    
    P1_0_Data = []
    P1_1_Data = []
    
    start_row = 2  # 2 because 1 is usually labels
    labels = dict()
    
    # Assigning labels to dictionary
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'A{row}']
        if cell.value is not None:
            labels[cell.value] = 0  # Append the value to the list
        else:
            break
            
    # P1 Data, Element 0
    for row in range(start_row, sheet.max_row + 1):  
            cell = sheet[f'{P1_0_Letter}{row}']
            forlabel = sheet[f'A{row}']
            if cell.value is not None:
                P1_0_Data.append(cell.value)  # Append the value to the list
                labels[forlabel.value] = cell.value
            else:
                # Stop the loop when an empty cell is encountered
                break
        
    # P1 Data, Element 1
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P1_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P1_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value
        else:
            # Stop the loop when an empty cell is encountered
            break
        
    solve = []
    
    if (isinstance(P2_0_Letter, float) or isinstance(P2_0_Letter, int)) and (isinstance(P2_1_Letter, float) or isinstance(P2_1_Letter, int)):
        # Answer
        for row in range(start_row, sheet.max_row + 1):
            for i in range(len(P1_0_Data)):
                dt1 = P1_0_Data[i] - P2_0_Letter
                dt2 =  P1_1_Data[i] - P2_1_Letter
                dt1 = pow(dt1, 2)
                dt2 = pow(dt2, 2)
                solve.append(m.sqrt(dt1 + dt2))
                labels[sheet[f'A{row}'].value] = m.sqrt(dt1 + dt2)
                row += 1
                
            workbook.close()
            break
        return solve, labels
        
    P2_0_Data = []
    P2_1_Data = []
        
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_0_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_0_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value
        else:
            # Stop the loop when an empty cell is encountered
            break
        
    # input("Second column for Second data: ") # e.g. B
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{P2_1_Letter}{row}']
        forlabel = sheet[f'A{row}']
        if cell.value is not None:
            P2_1_Data.append(cell.value)  # Append the value to the list
            labels[forlabel.value] = cell.value
        else:
            # Stop the loop when an empty cell is encountered
            break
     
    # Answer
    for row in range(start_row, sheet.max_row + 1):  
        for i in range(len(P1_0_Data)):
            dt1 = P1_0_Data[i] - P2_0_Data[i]
            dt2 =  P1_1_Data[i] - P2_1_Data[i]
            dt1 = pow(dt1, 2)
            dt2 = pow(dt2, 2)
            solve.append(m.sqrt(dt1 + dt2))
            labels[sheet[f'A{row}'].value] = m.sqrt(dt1 + dt2)
            row += 1
        workbook.close()
        break
    return solve, labels

"""
def manhattan(column_letter_1, column_letter_2):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
    sheet = workbook.active

    first_data = []
    second_data = []

    start_row = 2  # 2 because 1 is usually labels

    # Iterate through the column to get the first data
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_1}{row}']
        if cell.value is not None:
            first_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
        
    # input("Second column for Second data: ") # e.g. B
    for row in range(start_row, sheet.max_row + 1):  
        cell = sheet[f'{column_letter_2}{row}']
        if cell.value is not None:
            second_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break
    
    workbook.close()

    solve = []
    for i in range(len(first_data)):
        solve.append(first_data[i] - second_data[i])
    
    return solve
def minkowski(column_letter_1, column_letter_2, p):
    if p <= 0:
        raise ValueError("Parameter p must be greater than 0")
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook = xl.load_workbook(os.path.join(script_dir, 'data.xlsx'))
    sheet = workbook.active

    first_data = []
    second_data = []

    start_row = 2  # 2 because 1 is usually labels

    # Iterate through the column to get the first data
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f'{column_letter_1}{row}']
        if cell.value is not None:
            first_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break

    # Iterate through the second column for Second data
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f'{column_letter_2}{row}']
        if cell.value is not None:
            second_data.append(cell.value)  # Append the value to the list
        else:
            # Stop the loop when an empty cell is encountered
            break

    if len(first_data) != len(second_data):
        raise ValueError("Points must have the same number of elements/dimensions")

    workbook.close()
    
    solve = []
    for i in range(len(first_data)):
        d = abs(first_data[i] - second_data[i]) ** p
        solve.append(d ** (1 / p))
        
        ABS((B2-$D$3) + (C2-$E$3))**(1/2)
        
    return solve
"""

data, labels = euclidean('B', 'C', 'D', 'E', 'data')
# A and C for data 1, B and D for data 2
# Solve eucledian by: Point1(A,B) and Point2(C,D)

# data = manhattan('A', 'C')
# data = minkowski('A','B', 2)

goal_list = list(data)
sorted_labels = dict(sorted(labels.items(), key=lambda item: item[1])) # sorted by value
goal_list.sort()
root_node = build_tree(data, 0, len(labels)+1)

ranking = []
track = []

# Find the 5 lowest value using DLS
for i in range(len(labels)):
    ranking.append(depth_limited_search(root_node, goal_list[i], 0, len(labels)+1, track))

ranking_ = []
for sublist in ranking:
    if isinstance(sublist, list):
        for num in sublist:
            ranking_.append(num)
            continue
    ranking_.append(num)
        
store = set()
ranking__ = [x for x in ranking_ if x not in store and not store.add(x)]

print(f"Ranking of DLS: {ranking__[:len(labels)]}")

print("\nActual Ranking: ")
for name, val in sorted_labels.items():
    print(f"{name} : {val}")